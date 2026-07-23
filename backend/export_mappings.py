# Exports Carson's mapping spec out of the .xlsm into mappings.json.
#
#   python export_mappings.py "../reference/model 56.xlsm"
#   python apply_overrides.py        <- ALWAYS run this second
#
# Run this whenever a new model workbook lands. Nothing is transcribed by hand — the
# workbook is the source of truth and mappings.json is a build artifact of it.
#
# This script REBUILDS mappings.json from scratch, discarding whatever was there.
# Our own mapping decisions (the pre-NCI net income target, preferred stock
# issuance, the dropped Other-Income reclasses) are therefore NOT preserved by it —
# apply_overrides.py reapplies them, and skipping that step silently costs about
# half a point of reconcile pass rate with no error to warn you.
#
# The .xlsm is gitignored (an API key is embedded in it), so mappings.json is what
# actually ships: it carries the spec without carrying the secret.
#
# Only the *rules* are exported (model line, synonym, use type, priority). The
# Active / HasData / IsCandidate columns are deliberately NOT exported — those are
# live per-ticker formulas, and mapping_engine.py recomputes them from real FMP data
# at runtime. Freezing them would bake one company's shape into every other company:
# the workbook ships with PEP loaded, PEP has no R&D, so a frozen Active would zero
# out R&D for every company that does have it. Model 62's new "Applies now?" column
# on the reclass tables is the same kind of thing — `=IF($B115=$C$4,"● THIS TICKER","")`
# just highlights the loaded ticker's rows — so it is not exported either.
#
# Nothing about the sheet layout is hardcoded, because model 62 changed all of it:
# both tables moved one column right, and the reclass table now comes BEFORE the
# mapping dictionary instead of after it. Tables are located by their own header
# labels and bounded by each other, so either order and any column works.

import json
import sys
from pathlib import Path

import openpyxl

SHEETS = {
    "income_statement": "IS_Map",
    "balance_sheet": "BS_Map",
    "cash_flow": "CF_Map",
}

MASTER_COLS = ["statement", "section", "model_line", "synonym", "use_type", "priority", "active", "notes"]
RECLASS_COLS = ["ticker", "fiscal_year", "statement", "source_field", "from_line", "to_line", "amount", "reason"]

OUT_PATH = Path(__file__).resolve().parent / "mappings.json"


def _find_tables(rows):
    """Locate both tables by their own header labels, wherever they sit.

    Each header is matched on its first two labels rather than on a fixed column,
    and the row/column of each is returned, so a workbook that shifts columns or
    swaps the order of the two tables still exports.
    """
    master = reclass = None
    for i, r in enumerate(rows, start=1):
        vals = [str(v).strip() if v is not None else "" for v in r]
        for c, v in enumerate(vals):
            if master is None and v == "Statement" and vals[c + 1:c + 2] == ["Section"]:
                master = (i, c)
            if reclass is None and v == "Ticker" and vals[c + 1:c + 2] == ["Fiscal Year"]:
                reclass = (i, c)
    if master is None or reclass is None:
        raise SystemExit(f"could not find both tables (master={master}, reclass={reclass})")
    return master, reclass


def _body(rows, header_row, other_header_row):
    """The data rows of one table: everything below its header, stopping short of
    the other table's header. Either table may come first."""
    if other_header_row > header_row:
        return rows[header_row:other_header_row - 1]
    return rows[header_row:]


def export(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    out = {"source": Path(xlsm_path).name, "statements": {}}

    for statement, sheet_name in SHEETS.items():
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                 max_col=ws.max_column, values_only=True))
        (master_hdr, master_col), (reclass_hdr, reclass_col) = _find_tables(rows)

        master = []
        for r in _body(rows, master_hdr, reclass_hdr):
            cells = r[master_col:master_col + len(MASTER_COLS)]
            if not cells[2] or not cells[3]:  # a real mapping row needs a model line and a synonym
                continue
            row = dict(zip(MASTER_COLS, cells))
            row.pop("active")  # recomputed per ticker at runtime, never frozen
            row["priority"] = int(row["priority"]) if row["priority"] is not None else 99
            master.append(row)

        reclasses, skipped = [], 0
        for r in _body(rows, reclass_hdr, master_hdr):
            cells = r[reclass_col:reclass_col + len(RECLASS_COLS)]
            if not cells[0] or cells[6] in (None, ""):  # need a ticker and an amount
                continue
            row = dict(zip(RECLASS_COLS, cells))
            try:
                row["fiscal_year"] = int(row["fiscal_year"])
                row["amount"] = float(row["amount"])
            except (TypeError, ValueError):
                skipped += 1  # reported below rather than swallowed
                continue
            reclasses.append(row)
        if skipped:
            print(f"{sheet_name:<7} !! {skipped} reclass row(s) skipped: "
                  f"unreadable fiscal year or amount")

        out["statements"][statement] = {"master": master, "reclasses": reclasses}
        print(f"{sheet_name:<7} -> {statement:<17} {len(master):>3} mapping rows, {len(reclasses):>3} reclasses")

    OUT_PATH.write_text(json.dumps(out, indent=1))
    print(f"\nwrote {OUT_PATH}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit('usage: python export_mappings.py "<path to model.xlsm>"')
    export(sys.argv[1])
